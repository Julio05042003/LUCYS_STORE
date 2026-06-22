from django.shortcuts import render, redirect
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from apps.usuarios.decorators import rol_requerido
from apps.usuarios.helpers import *
from apps.usuarios.models import *
from django.db import transaction
from django.utils import timezone
from datetime import date
from django.utils.dateparse import parse_date
import re, json
from django.http import JsonResponse
from django.db import connection
from django.db.models import Q
from apps.usuarios.tokens import token_generator
from apps.caja.models import AperturaCaja
from apps.inventario.models import Inventario
from django.db.models.functions import TruncDate
from django.db.models import Sum, Count, F, DecimalField, ExpressionWrapper
from django.utils.timezone import now, timedelta
from decimal import Decimal
from apps.ventas.models import Venta, DetalleVenta
from apps.productos.models import Producto
from apps.inventario.models import Kardex, AjusteInventario
from apps.usuarios.models import Empleado
from apps.caja.models import MovimientoCaja
from django.contrib.auth.tokens import default_token_generator
from openpyxl import Workbook
from django.http import HttpResponse


# =========================
# LISTADO PRINCIPAL
# =========================
@login_required
def ubicaciones(request):

    sucursales = Sucursal.objects.select_related(
        'bodega',
        'direccion',
        'estado'
    ).prefetch_related(
        'telefonos__estado'
    ).all()

    for sucursal in sucursales:

        sucursal.telefonos_activos = sucursal.telefonos.filter(
            estado__nombre='Activo'
        )

    return render(
        request,
        "empleados/ubicaciones.html",
        {
            "sucursales": sucursales
        }
    )


# =========================
# CREAR SUCURSAL + BODEGA
# =========================
def crear_sucursal(request):
    if request.method == "POST":
        # 1. Captura de datos
        suc_nombre = request.POST.get("sucursal_nombre", "").strip()
        suc_codigo = request.POST.get("sucursal_codigo", "").strip()
        bod_nombre = request.POST.get("bodega_nombre", "").strip()
        bod_codigo = request.POST.get("bodega_codigo", "").strip()

        # 2. Validaciones de existencia (pre-transacción)
        if Sucursal.objects.filter(codigo=suc_codigo).exists():
            messages.error(request, f"Ya existe una sucursal con el código '{suc_codigo}'.")
            return redirect("ubicaciones")
        
        if Bodega.objects.filter(codigo=bod_codigo).exists():
            messages.error(request, f"Ya existe una bodega con el código '{bod_codigo}'.")
            return redirect("ubicaciones")

        try:
            with transaction.atomic():
                estado_activo = Estado.objects.get(nombre="Activo")

                # 3. Crear Dirección
                direccion = Direccion.objects.create(
                    pais=request.POST.get("pais", ""),
                    departamento=request.POST.get("departamento", ""),
                    ciudad=request.POST.get("ciudad", ""),
                    detalle=request.POST.get("detalle", "")
                )

                # 4. Crear Sucursal
                sucursal = Sucursal.objects.create(
                    nombre=suc_nombre,
                    codigo=suc_codigo,
                    direccion=direccion,
                    estado=estado_activo
                )

                # 5. Crear Bodega
                Bodega.objects.create(
                    nombre=bod_nombre,
                    codigo=bod_codigo,
                    direccion=direccion,
                    sucursal=sucursal,
                    estado=estado_activo
                )

                # 6. Procesar Teléfonos
                telefonos = request.POST.getlist("telefono[]")
                operadoras = request.POST.getlist("operadora[]")
                tipos = request.POST.getlist("tipo[]")
                
                # Zip permite recorrer las listas en paralelo
                contador_telefonos = 0
                for tel, op, tipo in zip(telefonos, operadoras, tipos):
                    tel_limpio = tel.strip().replace('-', '')
                    
                    if tel_limpio:
                        if not tel_limpio.isdigit():
                            raise ValueError(f"El teléfono '{tel}' debe ser numérico.")
                        
                        TelefonoSucursal.objects.create(
                            sucursal=sucursal,
                            estado=estado_activo,
                            numero=tel_limpio,
                            operadora=op,
                            tipo=tipo or "Oficina"
                        )
                        contador_telefonos += 1
                
                if contador_telefonos == 0:
                    raise ValueError("Debes agregar al menos un número de teléfono.")

            messages.success(request, "Sucursal, bodega y teléfonos creados correctamente.")
            
        except ValueError as ve:
            messages.error(request, str(ve))
        except Exception as e:
            messages.error(request, f"Error inesperado al procesar la creación: {str(e)}")
        
        return redirect("ubicaciones") 
    
    
# =========================
# EDITAR SUCURSAL + BODEGA
# =========================

@login_required
def editar_sucursal(request, id):
    sucursal = get_object_or_404(Sucursal, sucursal_id=id)

    if request.method != 'POST':
        return redirect('ubicaciones')

    try:
        with transaction.atomic():

            # =====================================
            # DATOS GENERALES
            # =====================================
            nuevo_nombre = request.POST.get('sucursal_nombre', '').strip()
            nuevo_codigo = request.POST.get('sucursal_codigo', '').strip()
            nueva_bodega = request.POST.get('bodega_nombre', '').strip()
            nuevo_cod_bodega = request.POST.get('bodega_codigo', '').strip()

            # =====================================
            # VALIDACIONES DE DUPLICADOS
            # =====================================
            if Sucursal.objects.filter(
                nombre__iexact=nuevo_nombre
            ).exclude(sucursal_id=id).exists():
                messages.error(
                    request,
                    f"Ya existe una sucursal llamada '{nuevo_nombre}'."
                )
                return redirect('ubicaciones')

            if Sucursal.objects.filter(
                codigo__iexact=nuevo_codigo
            ).exclude(sucursal_id=id).exists():
                messages.error(
                    request,
                    f"El código '{nuevo_codigo}' ya está asignado a otra sucursal."
                )
                return redirect('ubicaciones')

            if Bodega.objects.filter(
                nombre__iexact=nueva_bodega
            ).exclude(sucursal__sucursal_id=id).exists():
                messages.error(
                    request,
                    f"Ya existe una bodega llamada '{nueva_bodega}'."
                )
                return redirect('ubicaciones')

            if Bodega.objects.filter(
                codigo__iexact=nuevo_cod_bodega
            ).exclude(sucursal__sucursal_id=id).exists():
                messages.error(
                    request,
                    f"El código de bodega '{nuevo_cod_bodega}' ya está en uso."
                )
                return redirect('ubicaciones')

            # =====================================
            # ACTUALIZAR SUCURSAL
            # =====================================
            sucursal.nombre = nuevo_nombre
            sucursal.codigo = nuevo_codigo
            sucursal.save()

            # =====================================
            # ACTUALIZAR DIRECCIÓN
            # =====================================
            direccion = sucursal.direccion
            direccion.pais = request.POST.get('pais', '').strip()
            direccion.departamento = request.POST.get('departamento', '').strip()
            direccion.ciudad = request.POST.get('ciudad', '').strip()
            direccion.detalle = request.POST.get('detalle', '').strip()
            direccion.save()

            # =====================================
            # ACTUALIZAR BODEGA
            # =====================================
            bodega = sucursal.bodega
            bodega.nombre = nueva_bodega
            bodega.codigo = nuevo_cod_bodega
            bodega.save()

            # =====================================
            # TELÉFONOS
            # =====================================
            estado_activo = Estado.objects.get(nombre__iexact='Activo')
            estado_inactivo = Estado.objects.get(nombre__iexact='Inactivo')

            telefono_ids = request.POST.getlist('telefono_id[]')
            telefonos = request.POST.getlist('numero[]')
            operadoras = request.POST.getlist('operadora[]')
            tipos = request.POST.getlist('tipo[]')

            # Teléfonos activos actualmente en BD
            telefonos_actuales = {
                str(t.telefono_id): t
                for t in TelefonoSucursal.objects.filter(
                    sucursal=sucursal
                )
            }

            # IDs recibidos desde el formulario
            telefono_ids_recibidos = set()

            for i in range(len(telefonos)):

                telefono_id = (
                    telefono_ids[i].strip()
                    if i < len(telefono_ids)
                    else ''
                )

                numero = (
                    telefonos[i].strip().replace('-', '')
                    if i < len(telefonos)
                    else ''
                )

                operadora = (
                    operadoras[i].strip()
                    if i < len(operadoras)
                    else ''
                )

                tipo = (
                    tipos[i].strip()
                    if i < len(tipos)
                    else 'Ventas'
                )

                if not numero:
                    continue

                if not numero.isdigit():
                    raise ValueError(
                        f'El teléfono "{numero}" debe contener únicamente números.'
                    )

                # =====================================
                # NUEVO TELÉFONO
                # =====================================
                if telefono_id == '':

                    TelefonoSucursal.objects.create(
                        sucursal=sucursal,
                        numero=numero,
                        operadora=operadora,
                        tipo=tipo,
                        estado=estado_activo
                    )

                # =====================================
                # ACTUALIZAR TELÉFONO EXISTENTE
                # =====================================
                else:

                    telefono_ids_recibidos.add(telefono_id)

                    telefono = TelefonoSucursal.objects.get(
                        telefono_id=telefono_id,
                        sucursal=sucursal
                    )

                    telefono.numero = numero
                    telefono.operadora = operadora
                    telefono.tipo = tipo
                    telefono.estado = estado_activo
                    telefono.save()

            # =====================================
            # DESACTIVAR LOS ELIMINADOS
            # =====================================
            for telefono_id, telefono in telefonos_actuales.items():

                if telefono_id not in telefono_ids_recibidos:

                    telefono.estado = estado_inactivo
                    telefono.save()

            messages.success(
                request,
                'Ubicación actualizada correctamente.'
            )

            return redirect('ubicaciones')

    except Exception as e:
        messages.error(
            request,
            f'Error al actualizar: {str(e)}'
        )
        return redirect('ubicaciones')

# =========================
# CAMBIAR ESTADO (ACTIVO/INACTIVO)
# =========================

def cambiar_estado_sucursal(request, id):
    sucursal = get_object_or_404(Sucursal, sucursal_id=id)
    
    # Obtenemos los objetos de estado por su nombre
    estado_activo = Estado.objects.get(nombre="Activo")
    estado_inactivo = Estado.objects.get(nombre="Inactivo")

    # SI QUIERE ACTIVAR LA SUCURSAL
    if sucursal.estado == estado_inactivo:
        sucursal.estado = estado_activo
        sucursal.save()
        if hasattr(sucursal, "bodega"):
            sucursal.bodega.estado = estado_activo
            sucursal.bodega.save()
        messages.success(request, "Sucursal reactivada correctamente.")
        return redirect("ubicaciones")

    # SI QUIERE INACTIVAR (VALIDACIONES)
    
    # 1. Validar cajas abiertas: 
    # Buscamos aperturas de cajas asociadas a las cajas de la sucursal donde fecha_cierre es nulo
    cajas_abiertas = AperturaCaja.objects.filter(
        caja__sucursal=sucursal, 
        fecha_cierre__isnull=True
    ).exists()
    
    if cajas_abiertas:
        messages.error(request, "No se puede inactivar: existen cajas abiertas en esta sucursal.")
        return redirect("ubicaciones")

    # 2. Validar inventario en 0:
    # Si la sucursal tiene bodega, sumamos el stock de todos sus registros en inventario
    if hasattr(sucursal, "bodega"):
        tiene_stock = Inventario.objects.filter(
            bodega=sucursal.bodega, 
            stock__gt=0
        ).exists()
        
        if tiene_stock:
            messages.error(request, "No se puede inactivar: la bodega aún tiene productos en inventario.")
            return redirect("ubicaciones")

    # 3. PROCESO DE INACTIVACIÓN (con transacciones para integridad)
    with transaction.atomic():
        # Cambiar estados
        sucursal.estado = estado_inactivo
        sucursal.save()
        
        if hasattr(sucursal, "bodega"):
            sucursal.bodega.estado = estado_inactivo
            sucursal.bodega.save()

        # Quitar permisos de logueo a todos los empleados de la sucursal
        empleados = sucursal.empleado_set.all()
        for emp in empleados:
            if emp.user:
                emp.user.is_active = False
                emp.user.save()

    messages.success(request, "Sucursal inactivada exitosamente y accesos bloqueados.")
    return redirect("ubicaciones")


@login_required
def dashboard_view(request):
    usuario = request.user
    
    # 1. Control de Permisos por Rol / Sucursal
    try:
        empleado_perfil = usuario.empleado
        rol_usuario = empleado_perfil.rol.nombre.lower() # 'admin', 'gerente', etc.
        sucursal_usuario = empleado_perfil.sucursal
    except Empleado.DoesNotExist:
        # En caso de que sea un superusuario de Django sin perfil de empleado
        rol_usuario = 'administrador'
        sucursal_usuario = None

    # Base de querysets filtrados por seguridad
    ventas_qs = Venta.objects.filter(estado__nombre__iexact='Pagada')
    detalles_qs = DetalleVenta.objects.filter(venta__estado__nombre__iexact='Pagada')
    ajustes_qs = AjusteInventario.objects.filter(tipo='SALIDA')
    
    # Si es Gerente, limitar estrictamente a su sucursal
    if rol_usuario == 'gerente' and sucursal_usuario:
        ventas_qs = ventas_qs.filter(empleado__sucursal=sucursal_usuario)
        detalles_qs = detalles_qs.filter(venta__empleado__sucursal=sucursal_usuario)
        ajustes_qs = ajustes_qs.filter(empleado__sucursal=sucursal_usuario)
    elif rol_usuario == 'administrador':
        # Si es admin y seleccionó una sucursal específica en el filtro HTML
        sucursal_id = request.GET.get('sucursal')

        # 🔥 FIX AGREGADO
        if sucursal_id in [None, '', 'None']:
            sucursal_id = None

        if sucursal_id:
            ventas_qs = ventas_qs.filter(empleado__sucursal_id=sucursal_id)
            detalles_qs = detalles_qs.filter(venta__empleado__sucursal_id=sucursal_id)
            ajustes_qs = ajustes_qs.filter(empleado__sucursal_id=sucursal_id)

    # 2. Filtros de Fecha (Por defecto: últimos 30 días)
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    

    # 🔥 FIX AGREGADO: limpiar "None"
    if fecha_inicio_str in [None, '', 'None']:
        fecha_inicio_str = None

    if fecha_fin_str in [None, '', 'None']:
        fecha_fin_str = None
        
    
    
    # 🔥 FIX AGREGADO: evitar parse_date(None)
    if fecha_inicio_str:
        fecha_inicio_parsed = parse_date(fecha_inicio_str)
        ventas_qs = ventas_qs.filter(fecha__date__gte=fecha_inicio_parsed)
        detalles_qs = detalles_qs.filter(venta__fecha__date__gte=fecha_inicio_parsed)
        ajustes_qs = ajustes_qs.filter(fecha__date__gte=fecha_inicio_parsed)

    if fecha_fin_str:
        fecha_fin_parsed = parse_date(fecha_fin_str)
        ventas_qs = ventas_qs.filter(fecha__date__lte=fecha_fin_parsed)
        detalles_qs = detalles_qs.filter(venta__fecha__date__lte=fecha_fin_parsed)
        ajustes_qs = ajustes_qs.filter(fecha__date__lte=fecha_fin_parsed)
        
    hoy = date.today()    
    # CONVERTIR A DATE SEGURO
    fecha_inicio = parse_date(fecha_inicio_str) if fecha_inicio_str else None
    fecha_fin = parse_date(fecha_fin_str) if fecha_fin_str else None

    # DEFAULTS
    if not fecha_inicio:
        fecha_inicio = hoy.replace(day=1)

    if not fecha_fin:
        fecha_fin = hoy

    # VALIDACIÓN CORRECTA (YA COMO DATE, NO STRING)

    if fecha_fin < fecha_inicio:
        messages.error(request, "La fecha final no puede ser menor que la fecha inicial.")
        fecha_inicio, fecha_fin = fecha_fin, fecha_inicio

    # 3. Métrica: Ventas Totales y cantidad de transacciones
    metricas_generales = ventas_qs.aggregate(
        total_dinero=Sum('total'),
        total_transacciones=Count('venta_id')
    )
    total_ventas = metricas_generales['total_dinero'] or 0
    cantidad_ventas = metricas_generales['total_transacciones'] or 0
    
    # Metrica: Costo de ventas
    costo_venta = detalles_qs.aggregate(
        total=Sum(
            ExpressionWrapper(
                F('cantidad') * F('producto__precio_c'),
                output_field=DecimalField(
                    max_digits=15,
                    decimal_places=2
                )
            )
        )
    )['total'] or 0
    
    # Metrica de utilidad bruta
    utilidad_bruta = total_ventas - costo_venta

    margen_bruto = 0

    if total_ventas > 0:
        margen_bruto = round(
            (utilidad_bruta / total_ventas) * 100,
            2
        )
        
    # Metrica: perdidas
    perdidas = ajustes_qs.filter(
        motivo__in=[
            'PRODUCTO_DAÑADO',
            'PERDIDA'
        ]
    ).aggregate(
        total=Sum(
            ExpressionWrapper(
                F('cantidad') * F('producto__precio_c'),
                output_field=DecimalField(
                    max_digits=15,
                    decimal_places=2
                )
            )
        )
    )['total'] or 0
    
    #metrica regalias
    regalias = ajustes_qs.filter(
        motivo='REGALIA'
    ).aggregate(
        total=Sum(
            ExpressionWrapper(
                F('cantidad') * F('producto__precio_c'),
                output_field=DecimalField(
                    max_digits=15,
                    decimal_places=2
                )
            )
        )
    )['total'] or 0

    # 4. Métrica: Ventas por Empleado
    ventas_empleado_qs = ventas_qs.values(
        'empleado__user__first_name', 'empleado__user__last_name', 'empleado__user__username'
    ).annotate(
        total_vendido=Sum('total')
    ).order_by('-total_vendido')

    empleados_labels = []
    empleados_data = []
    for emp in ventas_empleado_qs:
        nombre = f"{emp['empleado__user__first_name']} {emp['empleado__user__last_name']}".strip() or emp['empleado__user__username']
        empleados_labels.append(nombre)
        empleados_data.append(float(emp['total_vendido']))
        

    # 5. Métrica: Productos Más Vendidos (Top 10 con su Categoría y Marca)
    productos_mas_vendidos = detalles_qs.values(
        'producto__nombre', 'producto__categoria__nombre', 'producto__marca__nombre'
    ).annotate(
        cantidad_total=Sum('cantidad'),
        ingreso_total=Sum('precio')
    ).filter(
        cantidad_total__gte=18
    ).order_by('-cantidad_total')[:10]

    # Datos estructurados para gráficos de torta/barras de categorías/marcas más vendidas
    cat_data = detalles_qs.values('producto__categoria__nombre').annotate(total=Sum('cantidad')).order_by('-total')[:5]
    categorias_labels = [c['producto__categoria__nombre'] for c in cat_data]
    categorias_valores = [c['total'] for c in cat_data]

    # 6. Métrica: Productos Menos Vendidos (Flop 10)
    productos_menos_vendidos = detalles_qs.values(
    'producto__nombre', 'producto__codigo', 'producto__categoria__nombre'
).annotate(
    cantidad_total=Sum('cantidad')
).filter(
    cantidad_total__lt=18
).order_by('cantidad_total')[:10]

    # Lista de sucursales para el combo de filtro (solo útil para el Admin)
    sucursales = Sucursal.objects.all() if rol_usuario == 'administrador' else None

    context = {
        'rol_usuario': rol_usuario,
        'sucursal_usuario': sucursal_usuario,
        'sucursales': sucursales,
        'total_ventas': total_ventas,
        'costo_venta': costo_venta,
        'utilidad_bruta': utilidad_bruta,
        'perdidas': perdidas,
        'regalias': regalias,
        'cantidad_ventas': cantidad_ventas,
        'productos_mas_vendidos': productos_mas_vendidos,
        'productos_menos_vendidos': productos_menos_vendidos,
        'chart_empleados_labels': json.dumps(empleados_labels),
        'chart_empleados_data': json.dumps(empleados_data),
        'chart_categorias_labels': json.dumps(categorias_labels),
        'chart_categorias_valores': json.dumps(categorias_valores),
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str,
        'sucursal_seleccionada': request.GET.get('sucursal', '')
    }

    return render(request, 'empleados/dashboard.html', context)

@login_required
def dashboard_excel(request):

    from datetime import date
    from django.utils.dateparse import parse_date
    from django.db.models import Sum, F, Count
    from openpyxl import Workbook

    usuario = request.user

    try:
        empleado = usuario.empleado
        rol_usuario = empleado.rol.nombre.lower()
        sucursal_usuario = empleado.sucursal
    except:
        rol_usuario = 'administrador'
        sucursal_usuario = None

    ventas_qs = Venta.objects.filter(estado__nombre__iexact='Pagada')
    detalles_qs = DetalleVenta.objects.filter(venta__estado__nombre__iexact='Pagada')
    ajustes_qs = AjusteInventario.objects.filter(tipo='SALIDA')

    # =====================
    # SUCURSAL
    # =====================

    sucursal_id = request.GET.get('sucursal')

    if sucursal_id in [None, '', 'None']:
        sucursal_id = None

    if rol_usuario == 'gerente' and sucursal_usuario:

        sucursal_nombre = sucursal_usuario.nombre

        ventas_qs = ventas_qs.filter(empleado__sucursal=sucursal_usuario)
        detalles_qs = detalles_qs.filter(venta__empleado__sucursal=sucursal_usuario)
        ajustes_qs = ajustes_qs.filter(empleado__sucursal=sucursal_usuario)

    elif rol_usuario == 'administrador' and sucursal_id:

        ventas_qs = ventas_qs.filter(empleado__sucursal_id=sucursal_id)
        detalles_qs = detalles_qs.filter(venta__empleado__sucursal_id=sucursal_id)
        ajustes_qs = ajustes_qs.filter(empleado__sucursal_id=sucursal_id)

        sucursal_nombre = f"Sucursal ID {sucursal_id}"
    else:
        sucursal_nombre = "Todas las sucursales"

    # =====================
    # FECHAS
    # =====================

    hoy = date.today()

    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    if fecha_inicio_str in [None, '', 'None']:
        fecha_inicio = hoy.replace(day=1)
    else:
        fecha_inicio = parse_date(fecha_inicio_str)

    if fecha_fin_str in [None, '', 'None']:
        fecha_fin = hoy
    else:
        fecha_fin = parse_date(fecha_fin_str)

    ventas_qs = ventas_qs.filter(fecha__date__range=(fecha_inicio, fecha_fin))
    detalles_qs = detalles_qs.filter(venta__fecha__date__range=(fecha_inicio, fecha_fin))
    ajustes_qs = ajustes_qs.filter(fecha__date__range=(fecha_inicio, fecha_fin))

    # =====================
    # KPIs
    # =====================

    total_ventas = ventas_qs.aggregate(total=Sum('total'))['total'] or 0
    total_transacciones = ventas_qs.count()

    costo_venta = detalles_qs.aggregate(
        total=Sum(F('cantidad') * F('producto__precio_c'))
    )['total'] or 0

    utilidad = total_ventas - costo_venta

    margen = (utilidad / total_ventas * 100) if total_ventas > 0 else 0

    ticket_promedio = (total_ventas / total_transacciones) if total_transacciones > 0 else 0

    # =====================
    # AJUSTES (CORREGIDO SIN ExpressionWrapper)
    # =====================

    ajustes_con_costo = ajustes_qs.annotate(
        costo_linea=F('cantidad') * F('producto__precio_c')
    )

    perdidas = ajustes_con_costo.filter(
        motivo__in=['PRODUCTO_DAÑADO', 'PERDIDA']
    ).aggregate(
        total=Sum('costo_linea')
    )['total'] or 0

    regalias = ajustes_con_costo.filter(
        motivo='REGALIA'
    ).aggregate(
        total=Sum('costo_linea')
    )['total'] or 0

    # =====================
    # EXCEL
    # =====================

    wb = Workbook()

    # =====================
    # HOJA 1
    # =====================
    ws1 = wb.active
    ws1.title = "Resumen"

    ws1.append(["REPORTE EJECUTIVO"])
    ws1.append(["Sucursal", sucursal_nombre])
    ws1.append(["Fecha inicio", fecha_inicio])
    ws1.append(["Fecha fin", fecha_fin])
    ws1.append([])
    ws1.append(["Ventas Totales", float(total_ventas)])
    ws1.append(["Transacciones", total_transacciones])
    ws1.append(["Ticket Promedio", float(ticket_promedio)])
    ws1.append(["Costo Venta", float(costo_venta)])
    ws1.append(["Utilidad", float(utilidad)])
    ws1.append(["Margen %", round(margen, 2)])
    ws1.append(["Pérdidas", float(perdidas)])
    ws1.append(["Regalías", float(regalias)])

    # =====================
    # HOJA 2
    # =====================
    ws2 = wb.create_sheet("Ventas Detalladas")

    ventas_det = ventas_qs.select_related('empleado', 'cliente')

    ws2.append(["Fecha", "Cliente", "Empleado", "Sucursal", "Total"])

    for v in ventas_det:
        ws2.append([
            v.fecha,
            getattr(v.cliente, 'nombre', 'N/A'),
            v.empleado.user.username,
            v.empleado.sucursal.nombre,
            float(v.total)
        ])

    # =====================
    # HOJA 3
    # =====================
    ws3 = wb.create_sheet("Ventas Empleados")

    ventas_emp = ventas_qs.values(
        'empleado__user__username'
    ).annotate(total=Sum('total')).order_by('-total')

    ws3.append(["Empleado", "Total", "% Participación"])

    for e in ventas_emp:
        porcentaje = (e['total'] / total_ventas * 100) if total_ventas > 0 else 0
        ws3.append([
            e['empleado__user__username'],
            float(e['total']),
            round(porcentaje, 2)
        ])

    # =====================
    # HOJA 4
    # =====================
    ws4 = wb.create_sheet("Productos Vendidos")

    productos = detalles_qs.values(
        'producto__nombre',
        'producto__categoria__nombre',
        'producto__marca__nombre'
    ).annotate(
        cantidad=Sum('cantidad'),
        ingreso=Sum('precio')
    ).order_by('-cantidad')

    ws4.append(["Producto", "Categoría", "Marca", "Cantidad", "Ingreso"])

    for p in productos:
        ws4.append([
            p['producto__nombre'],
            p['producto__categoria__nombre'],
            p['producto__marca__nombre'],
            p['cantidad'],
            float(p['ingreso'] or 0)
        ])

    # =====================
    # HOJA 5
    # =====================
    ws5 = wb.create_sheet("Ajustes")

    ajustes = ajustes_con_costo.values(
        'motivo',
        'producto__nombre'
    ).annotate(
        cantidad=Sum('cantidad'),
        costo=Sum('costo_linea')
    )

    ws5.append(["Motivo", "Producto", "Cantidad", "Costo"])

    for a in ajustes:
        ws5.append([
            a['motivo'],
            a['producto__nombre'],
            a['cantidad'],
            float(a['costo'] or 0)
        ])

    # =====================
    # RESPONSE
    # =====================

    from django.http import HttpResponse

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=REPORTE_ERP_DETALLADO.xlsx'

    wb.save(response)
    return response


def login_view(request):

    if request.method == 'POST':

        username = request.POST.get('usuario', '').strip()
        password = request.POST.get('password', '').strip()

        # VALIDAR CAMPOS
        if not username or not password:
            messages.error(request,"Todos los campos son obligatorios")
            return render(request, 'login.html')

        # VALIDAR USUARIO EXISTE
        try:
            user_db = User.objects.get(
                username=username
            )

        except User.DoesNotExist:
            messages.error(request,"El usuario no existe")
            return render(request, 'login.html')

        # INTENTOS
        intentos = request.session.get(
            f'intentos_{username}',
            0
        )

        # AUTENTICAR
        user = authenticate(
            request,
            username=username,
            password=password
        )

        # PASSWORD INCORRECTA
        if not user:

            # VALIDAR SI ES EMPLEADO
            es_empleado = Empleado.objects.filter(
                user=user_db
            ).exists()

            # SOLO EMPLEADOS SE BLOQUEAN
            if es_empleado:

                intentos += 1

                request.session[
                    f'intentos_{username}'
                ] = intentos

                if intentos >= 3:

                    user_db.is_active = False
                    user_db.save()

                    messages.error(
                        request,
                        "Cuenta bloqueada. Contacte al gerente o administrador."
                    )

                else:

                    messages.error(
                        request,
                        f"Credenciales incorrectas ({intentos}/3)"
                    )

            else:

                messages.error(
                    request,
                    "Correo o contraseña incorrectos"
                )

            return render(request, 'login.html')

        # RESET INTENTOS
        request.session[
            f'intentos_{username}'
        ] = 0

        # VALIDAR CLIENTE VERIFICADO
        es_empleado = Empleado.objects.filter(
            user=user
        ).exists()

        if not es_empleado and not user.is_active:

            messages.warning(
                request,
                "Debes verificar tu correo antes de iniciar sesión"
            )

            return render(request, 'login.html')

        # EMPLEADOS
        if es_empleado:

            empleado = Empleado.objects.select_related(
                'estado',
                'rol',
                'sucursal'
            ).get(user=user)

            # EMPLEADO BLOQUEADO
            if not user.is_active:

                messages.error(
                    request,
                    "Usuario bloqueado"
                )

                return render(request, 'login.html')

            # EMPLEADO INACTIVO
            if empleado.estado.nombre.lower() != "activo":

                messages.error(
                    request,
                    "Empleado inactivo"
                )

                return render(request, 'login.html')

            login(request, user)

            request.session['rol'] = empleado.rol.nombre
            request.session['nombre'] = user.first_name
            request.session['apellido'] = user.last_name

            return redirect('index_empleados')

        # CLIENTES
        login(request, user)

        return redirect('tienda')

    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login')



def password_reset_request(request):

    if request.method == 'POST':

        correo = request.POST.get('email')

        try:

            usuario = User.objects.get(email=correo)

            # Verificar si pertenece a un empleado
            empleado = Empleado.objects.filter(
                user=usuario
            ).select_related('rol').first()

            if empleado:

                rol = empleado.rol.nombre.upper().strip()

                # Solo ADMIN puede recuperar contraseña
                if rol != 'ADMIN':

                    messages.error(
                        request,
                        'Los empleados deben solicitar el cambio de contraseña al administrador.'
                    )

                    return redirect('password_reset')

            token = default_token_generator.make_token(usuario)

            url = request.build_absolute_uri(
                reverse(
                    'password_reset_confirm',
                    args=[usuario.id, token]
                )
            )

            asunto = 'Recuperación de contraseña'

            mensaje = f'''
Hola {usuario.first_name} {usuario.last_name},

Hemos recibido una solicitud para restablecer la contraseña de tu cuenta.

Haz clic en el siguiente enlace:

{url}

Si no realizaste esta solicitud, puedes ignorar este mensaje.
'''

            send_mail(
                asunto,
                mensaje,
                settings.DEFAULT_FROM_EMAIL,
                [usuario.email],
                fail_silently=False
            )

            messages.success(
                request,
                'Se ha enviado un enlace de recuperación a tu correo.'
            )

            return redirect('login')

        except User.DoesNotExist:

            messages.error(
                request,
                'No existe una cuenta asociada a ese correo.'
            )

    return render(
        request,
        'empleados/password_reset.html'
    )
    
     
#Vista para cambiar contraseña
def password_reset_confirm(request, user_id, token):

    try:
        usuario = User.objects.get(id=user_id)

    except User.DoesNotExist:
        messages.error(request, 'Usuario no válido.')
        return redirect('login')

    if not default_token_generator.check_token(usuario, token):
        messages.error(request, 'El enlace ha expirado o es inválido.')
        return redirect('login')

    if request.method == 'POST':

        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        if password1 != password2:

            messages.error(
                request,
                'Las contraseñas no coinciden.'
            )

            return render(
                request,
                'empleados/password_reset_confirm.html',
                {'usuario': usuario}
            )

        if not validar_password(password1):

            messages.error(
                request,
                'La contraseña debe tener al menos 8 caracteres, una letra mayúscula, una letra minúscula, un número y un carácter especial.'
            )

            return render(
                request,
                'empleados/password_reset_confirm.html',
                {'usuario': usuario}
            )

        usuario.set_password(password1)
        usuario.save()

        messages.success(
            request,
            'Contraseña actualizada correctamente.'
        )

        return redirect('login')

    return render(
        request,
        'empleados/password_reset_confirm.html',
        {'usuario': usuario}
    )
    
    
# =========================
# INDEX EMPLEADOS
# =========================
@login_required(login_url='login')
def index_empleados(request):

    try:
        empleado = Empleado.objects.select_related('user', 'rol', 'sucursal').get(user=request.user)
    except Empleado.DoesNotExist:
        return redirect('tienda')

    return render(request, 'empleados/index.html', {
        'empleado': empleado
    })


@login_required
@rol_requerido(['Administrador', 'Gerente'])
def usuarios_view(request):

    empleado_actual = request.empleado

    es_admin = empleado_actual.rol.nombre.lower() == "administrador"

    if es_admin:

        empleados = Empleado.objects.select_related(
            'user',
            'rol',
            'sucursal',
            'estado'
        ).all()

        roles = Rol.objects.all()

    else:

        empleados = Empleado.objects.select_related(
            'user',
            'rol',
            'sucursal',
            'estado'
        ).filter(
            sucursal=empleado_actual.sucursal
        )

        roles = Rol.objects.exclude(
            nombre__icontains="administrador"
        ).exclude(
            nombre__icontains="gerente"
        )

    sucursales = Sucursal.objects.all()

    estados = Estado.objects.filter(
        nombre__in=['Activo', 'Inactivo']
    )

    return render(request, 'empleados/usuarios.html', {

        'empleados': empleados,
        'roles': roles,
        'sucursales': sucursales,
        'estados': estados,
        'es_admin': es_admin

    })
    
    

def render_usuarios(request, extra_context=None):

    empleado_actual = request.empleado

    empleados = Empleado.objects.select_related(
        'user',
        'rol',
        'sucursal'
    )

    # gerente solo empleados de su sucursal
    if empleado_actual.rol.nombre.lower() != 'administrador':

        empleados = empleados.filter(
            sucursal_id=empleado_actual.sucursal_id
        )

    context = {

        'empleados': empleados,
        'roles': Rol.objects.all(),
        'sucursales': Sucursal.objects.all(),

    }

    if extra_context:
        context.update(extra_context)

    return render(
        request,
        'empleados/usuarios.html',
        context
    )

# =========================
# CREAR USUARIO
# =========================
@login_required
@rol_requerido(['Administrador', 'Gerente'])
def crear_usuario(request):

    if request.method != 'POST':
        return redirect('usuarios')

    empleado_actual = request.empleado

    email = request.POST.get('email', '').strip()
    username = request.POST.get('username', '').strip()
    password = request.POST.get('password', '').strip()

    if not validar_email(email):
        messages.error(request,"Correo inválido")
        return render_usuarios(
            request,
            {
                'abrir_modal': 'modalUsuario'
            }
        )

    if not validar_password(password):
        messages.error(request,"La contraseña debe tener mínimo 8 caracteres, mayúscula, minúscula, número y carácter especial")
        return render_usuarios(request,{'abrir_modal': 'modalUsuario'})

    if User.objects.filter(email=email).exists():
        messages.error(request,"El email ya existe")
        return render_usuarios(request,{'abrir_modal': 'modalUsuario'})
    
    if User.objects.filter(username=username).exists():
        messages.error(request,"El username ya existe")
        return render_usuarios(request,{'abrir_modal': 'modalUsuario'})

    if empleado_actual.rol.nombre.lower() == 'administrador': 
        sucursal_id = request.POST.get('sucursal_id')

    else:
        sucursal_id = empleado_actual.sucursal_id


    estado_activo = Estado.objects.filter(
        nombre__iexact='Activo'
    ).first()

    if not estado_activo:
        messages.error(request,"No existe el estado activo")
        return redirect('usuarios')

    # CREAR USER
    user = User.objects.create_user(

        username=username,
        email=email,
        password=password,

        first_name=request.POST.get('nombre'),
        last_name=request.POST.get('apellido')

    )

    Empleado.objects.create(
        user=user,
        rol_id=request.POST.get('rol_id'),
        sucursal_id=sucursal_id,
        estado=estado_activo
    )

    messages.success(request,"Usuario creado correctamente")
    return redirect('usuarios')


# =========================
# BLOQUEAR USUARIO
# =========================
@login_required
@rol_requerido(['Administrador', 'Gerente'])
def bloquear_usuario(request, user_id):

    empleado_actual = get_object_or_404(
        Empleado.objects.select_related('rol', 'sucursal'),
        user=request.user
    )

    user = get_object_or_404(User, id=user_id)

    # evitar auto bloqueo
    if request.user.id == user.id:
        messages.warning(request,"No puedes bloquearte a ti mismo")
        return redirect('usuarios')

    # validar que exista empleado
    try:

        empleado_objetivo = Empleado.objects.select_related(
            'rol',
            'sucursal'
        ).get(user=user)

    except Empleado.DoesNotExist:
        messages.error(request,"El usuario no pertenece a empleados")
        return redirect('usuarios')

    # VALIDAR PERMISOS POR SUCURSAL
    # administrador puede bloquear cualquiera
    if empleado_actual.rol.nombre.lower() != 'administrador':

        # gerente solo empleados de SU sucursal
        if empleado_objetivo.sucursal_id != empleado_actual.sucursal_id:

            messages.error(request,"No puedes bloquear empleados de otra sucursal")
            return redirect('usuarios')

        # gerente NO puede bloquear administradores
        if empleado_objetivo.rol.nombre.lower() == 'administrador':

            messages.error(request,"No puedes bloquear administradores")
            return redirect('usuarios')

    # BLOQUEAR
    user.is_active = False
    user.save()
    
    # opcional:
    # empleado_objetivo.estado = Estado.objects.get(nombre__iexact='Inactivo')
    # empleado_objetivo.save()

    messages.success(request,"Usuario bloqueado correctamente")
    return redirect('usuarios')


# =========================
# DESBLOQUEAR USUARIO
# =========================
@login_required
@rol_requerido(['Administrador', 'Gerente'])
def desbloquear_usuario(request, user_id):

    empleado_actual = get_object_or_404(
        Empleado.objects.select_related('rol', 'sucursal'),
        user=request.user
    )

    user = get_object_or_404(User, id=user_id)

    # validar empleado
    try:

        empleado_objetivo = Empleado.objects.select_related(
            'rol',
            'sucursal'
        ).get(user=user)

    except Empleado.DoesNotExist:
        messages.error(request,"El usuario no pertenece a empleados")
        return redirect('usuarios')

    # VALIDAR PERMISOS
    if empleado_actual.rol.nombre.lower() != 'administrador':

        # gerente solo SU sucursal
        if empleado_objetivo.sucursal_id != empleado_actual.sucursal_id:
            messages.error(request,"No puedes desbloquear empleados de otra sucursal")
            return redirect('usuarios')

        # gerente no desbloquea admins
        if empleado_objetivo.rol.nombre.lower() == 'administrador':
            messages.error(request,"No puedes desbloquear administradores")
            return redirect('usuarios')

    # DESBLOQUEAR
    user.is_active = True
    user.save()

    # opcional:
    # empleado_objetivo.estado = Estado.objects.get(nombre__iexact='Activo')
    # empleado_objetivo.save()

    messages.success(request,"Usuario desbloqueado correctamente")
    return redirect('usuarios')

# =========================
# EDITAR USUARIO
# =========================
@login_required
@rol_requerido(['Administrador', 'Gerente'])
def editar_usuario(request):

    if request.method != "POST":
        return redirect('usuarios')

    empleado_actual = request.empleado

    user_id = request.POST.get('user_id')

    user = get_object_or_404(User,id=user_id)
    emp = get_object_or_404(
        Empleado.objects.select_related(
            'rol',
            'sucursal',
            'user'
        ),
        user=user
    )

    # VALIDAR ACCESO GERENTE
    if empleado_actual.rol.nombre.lower() != 'administrador':
        # gerente solo puede editar empleados de su sucursal
        if emp.sucursal_id != empleado_actual.sucursal_id:
            messages.error(request,"No puedes editar empleados de otra sucursal")
            return redirect('usuarios')

    # DATOS
    email = request.POST.get('email','').strip()
    username = request.POST.get('username','').strip()
    nombre = request.POST.get('nombre','').strip()
    apellido = request.POST.get('apellido','').strip()

    # VALIDAR EMAIL
    if not validar_email(email):
        messages.error(request,"Correo inválido")
        return render_usuarios(request,{'abrir_modal': 'modalEditarUsuario'})

    # DUPLICADOS
    if User.objects.exclude(id=user_id).filter(email=email).exists():
        messages.error(request,"El email ya existe")
        return render_usuarios(request,{'abrir_modal': 'modalEditarUsuario'})

    if User.objects.exclude(id=user_id).filter(username=username).exists():
        messages.error(request,"El username ya existe")
        return render_usuarios(request,{'abrir_modal': 'modalEditarUsuario'})

    # ACTUALIZAR USER
    user.email = email
    user.username = username
    user.first_name = nombre
    user.last_name = apellido
    user.save()

    # ROL
    rol_id = request.POST.get('rol_id')

    # gerente NO puede asignar admin
    if empleado_actual.rol.nombre.lower() != 'administrador':
        rol_admin = Rol.objects.filter(nombre__icontains='administrador').first()
        rol_gerente = Rol.objects.filter(nombre__icontains='gerente').first()

        if str(rol_id) in [
            str(rol_admin.rol_id)
            if rol_admin else '',

            str(rol_gerente.rol_id)
            if rol_gerente else ''

        ]:
            messages.error(request,"No puedes asignar ese rol")
            return render_usuarios(request,{'abrir_modal': 'modalEditarUsuario'})

    # ACTUALIZAR EMPLEADO
    emp.rol_id = rol_id

    # administrador puede mover sucursal
    if empleado_actual.rol.nombre.lower() == 'administrador':
        emp.sucursal_id = request.POST.get('sucursal_id')
    else:
        # gerente mantiene su sucursal
        emp.sucursal_id = empleado_actual.sucursal_id

    emp.save()

    messages.success(request,"Usuario actualizado correctamente")
    return redirect('usuarios')

# =========================
# REGISTRO CLIENTE ONLINE
# =========================
def registro_cliente_view(request):

    if request.method != 'POST':
        return render(request, 'registro_cliente.html')

    try:

        with transaction.atomic():

            # DATOS PERSONALES
            email = request.POST.get('email', '').strip()
            password = request.POST.get('password', '').strip()

            nombre = request.POST.get('nombre', '').strip()
            apellido = request.POST.get('apellido', '').strip()
            identificacion = request.POST.get('identificacion', '').strip()

            # LISTAS DIRECCIONES
            paises = request.POST.getlist('pais[]')
            departamentos = request.POST.getlist('departamento[]')
            ciudades = request.POST.getlist('ciudad[]')
            detalles = request.POST.getlist('detalle[]')
            tipos_direccion = request.POST.getlist('tipo_direccion[]')

            # LISTAS TELÉFONOS
            telefonos = request.POST.getlist('telefono[]')
            operadoras = request.POST.getlist('operadora[]')
            tipos_tel = request.POST.getlist('tipo_tel[]')

            # VALIDACIONES
            if not all([
                email,
                password,
                nombre,
                apellido,
                identificacion
            ]):

                messages.error(request,"Todos los campos son obligatorios")
                return redirect('registro_cliente')

            if not validar_email(email):
                messages.error(request, "Correo inválido")
                return redirect('registro_cliente')

            if not validar_password(password):
                messages.error(request,"La contraseña debe tener mínimo 6 caracteres y un número")
                return redirect('registro_cliente')

            # username/email únicos
            if User.objects.filter(username=email).exists():
                messages.error(request, "El correo ya existe")
                return redirect('registro_cliente')

            if User.objects.filter(email=email).exists():
                messages.error(request, "El email ya existe")
                return redirect('registro_cliente')

            # identificación única
            if Cliente.objects.filter(
                identificacion=identificacion
            ).exists():

                messages.error(
                    request,
                    "La identificación ya existe"
                )

                return redirect('registro_cliente')

            # VALIDAR DIRECCIONES
            direcciones_validas = 0

            for i in range(len(paises)):

                pais = paises[i].strip()
                depto = departamentos[i].strip()
                ciudad = ciudades[i].strip()
                detalle = detalles[i].strip()

                if pais and depto and ciudad and detalle:
                    direcciones_validas += 1

            if direcciones_validas == 0:

                messages.error(request,"Debes agregar al menos una dirección")
                return redirect('registro_cliente')

            # VALIDAR TELÉFONOS
            telefonos_validos = 0

            for tel in telefonos:

                tel = tel.strip()

                if tel:

                    if not tel.isdigit():
                        messages.error(request,"Todos los teléfonos deben ser numéricos")
                        return redirect('registro_cliente')

                    telefonos_validos += 1

            if telefonos_validos == 0:
                messages.error(request,"Debes ingresar al menos un teléfono")
                return redirect('registro_cliente')

            # ESTADO ACTIVO
            estado_activo = Estado.objects.get(
                nombre__iexact="Activo"
            )

            # CREAR USER
            user = User.objects.create_user(
                username=email,
                email=email,
                password=password,
                first_name=nombre,
                last_name=apellido,
                is_active=False
            )

            # =========================
            # CREAR CLIENTE
            # =========================
            cliente = Cliente.objects.create(
                user=user,
                identificacion=identificacion,
                nombre=nombre,
                apellido=apellido
            )

            # =========================
            # DIRECCIONES
            # =========================
            for i in range(len(paises)):

                pais = paises[i].strip()
                depto = departamentos[i].strip()
                ciudad = ciudades[i].strip()
                detalle = detalles[i].strip()

                if pais and depto and ciudad and detalle:

                    direccion = Direccion.objects.create(
                        pais=pais,
                        departamento=depto,
                        ciudad=ciudad,
                        detalle=detalle
                    )

                    ClienteDireccion.objects.create(
                        cliente=cliente,
                        direccion=direccion,
                        estado=estado_activo,
                        tipo=tipos_direccion[i]
                        if i < len(tipos_direccion)
                        else "Casa"
                    )

            # =========================
            # TELÉFONOS
            # =========================
            for i in range(len(telefonos)):

                numero = telefonos[i].strip()

                if numero:

                    TelefonoCliente.objects.create(
                        cliente=cliente,
                        estado=estado_activo,
                        numero=numero,
                        operadora=operadoras[i]
                        if i < len(operadoras)
                        else "",
                        tipo=tipos_tel[i]
                        if i < len(tipos_tel)
                        else ""
                    )
                    
            # ENVIAR VERIFICACIÓN
            enviar_verificacion_email(request,user)
            
            messages.success(request,"Cuenta creada correctamente. Revisa tu correo para activar la cuenta.")

            return redirect('login')

    except Exception as e:
        messages.error(request,f"Error inesperado: {str(e)}")
        return redirect('registro_cliente')


def activar_cuenta(request, uid, token):

    try:
        user = User.objects.get(id=uid)

    except User.DoesNotExist:
        messages.error(request,"Usuario inválido")
        return redirect('login')

    if token_generator.check_token(user, token):
        user.is_active = True
        user.save()
        messages.success(request,"Cuenta verificada correctamente")
        return redirect('login')

    messages.error(request,"Token inválido o expirado")
    return redirect('login')

# =========================
# LISTAR CLIENTES
# =========================
@login_required
def clientes_view(request):

    clientes = Cliente.objects.select_related(
        'user'
    ).all().order_by('-cliente_id')

    return render(request, 'empleados/clientes.html', {
        'clientes': clientes
    })


# =========================
# CREAR CLIENTE FÍSICO
# =========================
@login_required
def crear_cliente(request):

    origen = request.POST.get('origen', 'clientes')

    def redireccion():

        if origen == 'ventas':
            return redirect('ventas')

        return redirect('clientes')

    if request.method != 'POST':
        return redireccion()

    try:

        with transaction.atomic():

            # =========================
            # DATOS PERSONALES
            # =========================
            nombre = request.POST.get(
                'nombre',
                ''
            ).strip()

            apellido = request.POST.get(
                'apellido',
                ''
            ).strip()

            identificacion = request.POST.get(
                'identificacion',
                ''
            ).strip()

            # LISTAS
            paises = request.POST.getlist('pais[]')
            departamentos = request.POST.getlist('departamento[]')
            ciudades = request.POST.getlist('ciudad[]')
            detalles = request.POST.getlist('detalle[]')
            tipos_direccion = request.POST.getlist('tipo_direccion[]')

            telefonos = request.POST.getlist('telefono[]')
            operadoras = request.POST.getlist('operadora[]')
            tipos_tel = request.POST.getlist('tipo_tel[]')

            # VALIDACIONES
            if not nombre or not apellido or not identificacion:
                messages.error(request, "Todos los campos son obligatorios")
                return redireccion()

            # IDENTIFICACIÓN ÚNICA
            if Cliente.objects.filter(
                identificacion=identificacion
            ).exists():
                messages.error(request,"La identificación ya existe")
                return redireccion()

            # VALIDAR TELÉFONOS
            telefonos_validos = []

            for tel in telefonos:

                tel = tel.strip()

                if tel:

                    telefono_limpio = tel.replace('-', '')

                    if not telefono_limpio.isdigit():
                        messages.error(request, "Los teléfonos deben ser numéricos")
                        return redireccion()

                    telefonos_validos.append(telefono_limpio)

            if len(telefonos_validos) == 0:
                messages.error(request, "Debes agregar al menos un teléfono")
                return redireccion()

            # ESTADO ACTIVO
            estado_activo = Estado.objects.filter(
                nombre__iexact='Activo'
            ).first()

            if not estado_activo:
                messages.error(request,"No existe el estado ACTIVO")
                return redireccion()

            # CREAR CLIENTE
            cliente = Cliente.objects.create(
                nombre=nombre,
                apellido=apellido,
                identificacion=identificacion
            )

            # GUARDAR DIRECCIONES
            total_direcciones = min(
                len(paises),
                len(departamentos),
                len(ciudades),
                len(detalles)
            )

            for i in range(total_direcciones):

                pais = paises[i].strip()
                depto = departamentos[i].strip()
                ciudad = ciudades[i].strip()
                detalle = detalles[i].strip()

                if not all([
                    pais,
                    depto,
                    ciudad,
                    detalle
                ]):
                    continue

                direccion = Direccion.objects.create(
                    pais=pais,
                    departamento=depto,
                    ciudad=ciudad,
                    detalle=detalle
                )

                ClienteDireccion.objects.create(
                    cliente=cliente,
                    direccion=direccion,
                    estado=estado_activo,
                    tipo=(
                        tipos_direccion[i]
                        if i < len(tipos_direccion)
                        else "Casa"
                    )
                )

            # GUARDAR TELÉFONOS
            for i in range(len(telefonos)):

                numero = telefonos[i].strip()

                if not numero:
                    continue

                numero = numero.replace('-', '')

                operadora = (
                    operadoras[i]
                    if i < len(operadoras)
                    else ""
                )

                tipo = (
                    tipos_tel[i]
                    if i < len(tipos_tel)
                    else "Personal"
                )

                TelefonoCliente.objects.create(
                    cliente=cliente,
                    estado=estado_activo,
                    numero=numero,
                    operadora=operadora,
                    tipo=tipo
                )

            messages.success(
                request,
                "Cliente físico creado correctamente"
            )

            return redireccion()

    except Exception as e:

        print("ERROR CREAR CLIENTE:", str(e))

        messages.error(
            request,
            f"Error: {str(e)}"
        )

        return redireccion()

# =========================
# EDITAR CLIENTE
# =========================
@login_required
def editar_cliente(request):

    if request.method != 'POST':
        return redirect('clientes')

    try:

        with transaction.atomic():

            cliente_id = request.POST.get('cliente_id')

            cliente = get_object_or_404(
                Cliente,
                cliente_id=cliente_id
            )

            nombre = request.POST.get(
                'nombre',
                ''
            ).strip()

            apellido = request.POST.get(
                'apellido',
                ''
            ).strip()

            identificacion = request.POST.get(
                'identificacion',
                ''
            ).strip()

            # =========================
            # VALIDACIONES
            # =========================

            if not all([
                nombre,
                apellido,
                identificacion
            ]):

                messages.error(
                    request,
                    "Todos los campos son obligatorios"
                )

                return redirect('clientes')

            existe = Cliente.objects.exclude(
                cliente_id=cliente.cliente_id
            ).filter(
                identificacion=identificacion
            ).exists()

            if existe:

                messages.error(
                    request,
                    "La identificación ya existe"
                )

                return redirect('clientes')

            # =========================
            # ACTUALIZAR CLIENTE
            # =========================

            cliente.nombre = nombre
            cliente.apellido = apellido
            cliente.identificacion = identificacion

            cliente.save()

            # =========================
            # ACTUALIZAR USER
            # =========================

            if cliente.user:

                cliente.user.first_name = nombre
                cliente.user.last_name = apellido

                cliente.user.save()

            # =========================
            # ESTADOS
            # =========================

            estado_inactivo = Estado.objects.get(
                nombre__iexact='Inactivo'
            )

            estado_activo = Estado.objects.get(
                nombre__iexact='Activo'
            )

            # ==================================================
            # TELÉFONOS
            # ==================================================

            telefono_ids = request.POST.getlist('telefono_id[]')
            telefonos = request.POST.getlist('telefono[]')
            operadoras = request.POST.getlist('operadora[]')
            tipos_tel = request.POST.getlist('tipo_tel[]')
            telefonos_eliminar = request.POST.getlist(
                'telefono_eliminar[]'
            )

            for i in range(len(telefonos)):

                telefono_id = (
                    telefono_ids[i].strip()
                    if i < len(telefono_ids)
                    else ''
                )

                numero = telefonos[i].strip()

                eliminar = (
                    telefonos_eliminar[i]
                    if i < len(telefonos_eliminar)
                    else '0'
                )

                # =========================
                # NUEVO VACÍO
                # =========================

                if not numero:

                    continue

                numero = numero.replace('-', '')

                # =========================
                # VALIDAR
                # =========================

                if not numero.isdigit():

                    messages.error(
                        request,
                        "Los teléfonos deben ser numéricos"
                    )

                    return redirect('clientes')

                operadora = (
                    operadoras[i]
                    if i < len(operadoras)
                    else ''
                )

                tipo = (
                    tipos_tel[i]
                    if i < len(tipos_tel)
                    else 'Personal'
                )

                # =========================
                # ACTUALIZAR EXISTENTE
                # =========================

                if telefono_id:

                    telefono = TelefonoCliente.objects.get(
                        telefono_id=telefono_id
                    )

                    # ELIMINAR
                    if eliminar == '1':

                        telefono.estado = estado_inactivo
                        telefono.save()

                        continue

                    telefono.numero = numero
                    telefono.operadora = operadora
                    telefono.tipo = tipo
                    telefono.estado = estado_activo

                    telefono.save()

                # =========================
                # CREAR NUEVO
                # =========================

                else:

                    TelefonoCliente.objects.create(
                        cliente=cliente,
                        estado=estado_activo,
                        numero=numero,
                        operadora=operadora,
                        tipo=tipo
                    )

            # ==================================================
            # DIRECCIONES
            # ==================================================

            direccion_ids = request.POST.getlist('direccion_id[]')

            paises = request.POST.getlist('pais[]')
            departamentos = request.POST.getlist('departamento[]')
            ciudades = request.POST.getlist('ciudad[]')
            detalles = request.POST.getlist('detalle[]')
            tipos_direccion = request.POST.getlist(
                'tipo_direccion[]'
            )

            direcciones_eliminar = request.POST.getlist(
                'direccion_eliminar[]'
            )

            for i in range(len(paises)):

                direccion_id = (
                    direccion_ids[i].strip()
                    if i < len(direccion_ids)
                    else ''
                )

                pais = paises[i].strip()
                depto = departamentos[i].strip()
                ciudad = ciudades[i].strip()
                detalle = detalles[i].strip()

                eliminar = (
                    direcciones_eliminar[i]
                    if i < len(direcciones_eliminar)
                    else '0'
                )

                # =========================
                # VALIDAR VACÍOS
                # =========================

                if not all([
                    pais,
                    depto,
                    ciudad,
                    detalle
                ]):

                    continue

                tipo = (
                    tipos_direccion[i]
                    if i < len(tipos_direccion)
                    else 'Casa'
                )

                # =========================
                # ACTUALIZAR EXISTENTE
                # =========================

                if direccion_id:

                    cliente_direccion = ClienteDireccion.objects.get(
                        cliente_direccion_id=direccion_id
                    )

                    # ELIMINAR
                    if eliminar == '1':

                        cliente_direccion.estado = estado_inactivo
                        cliente_direccion.save()

                        continue

                    direccion = cliente_direccion.direccion

                    direccion.pais = pais
                    direccion.departamento = depto
                    direccion.ciudad = ciudad
                    direccion.detalle = detalle

                    direccion.save()

                    cliente_direccion.tipo = tipo
                    cliente_direccion.estado = estado_activo

                    cliente_direccion.save()

                # =========================
                # CREAR NUEVA
                # =========================

                else:

                    direccion = Direccion.objects.create(
                        pais=pais,
                        departamento=depto,
                        ciudad=ciudad,
                        detalle=detalle
                    )

                    ClienteDireccion.objects.create(
                        cliente=cliente,
                        direccion=direccion,
                        estado=estado_activo,
                        tipo=tipo
                    )

            # =========================
            # SUCCESS
            # =========================

            messages.success(
                request,
                "Cliente actualizado correctamente"
            )

            return redirect('clientes')

    except Exception as e:

        print("ERROR EDITAR CLIENTE:", str(e))

        messages.error(
            request,
            f"Error: {str(e)}"
        )

        return redirect('clientes')

# =========================
# CREAR USER CLIENTE
# =========================
@login_required
def crear_user_cliente(request):

    if request.method != 'POST':
        return redirect('clientes')

    try:

        cliente_id = request.POST.get('cliente_id')

        cliente = get_object_or_404(
            Cliente,
            cliente_id=cliente_id
        )

        if cliente.user:

            messages.warning(
                request,
                "El cliente ya tiene cuenta"
            )

            return redirect('clientes')

        email = request.POST.get(
            'email',
            ''
        ).strip()

        username = request.POST.get(
            'username',
            ''
        ).strip()

        password = request.POST.get(
            'password',
            ''
        ).strip()

        # VALIDACIONES
        if not validar_email(email):

            messages.error(
                request,
                "Correo inválido"
            )

            return redirect('clientes')

        if not validar_password(password):

            messages.error(
                request,
                "Contraseña inválida"
            )

            return redirect('clientes')

        if User.objects.filter(
            username=username
        ).exists():

            messages.error(
                request,
                "El username ya existe"
            )

            return redirect('clientes')

        if User.objects.filter(
            email=email
        ).exists():

            messages.error(
                request,
                "El email ya existe"
            )

            return redirect('clientes')

        # CREAR USER
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=cliente.nombre,
            last_name=cliente.apellido
        )

        cliente.user = user
        cliente.save()

        messages.success(
            request,
            "Cuenta online creada correctamente"
        )

        return redirect('clientes')

    except Exception as e:

        messages.error(
            request,
            f"Error: {str(e)}"
        )

        return redirect('clientes')


# =========================
# PERFIL CLIENTE ONLINE
# =========================
@login_required
def perfil_cliente(request):

    try:

        cliente = get_object_or_404(
            Cliente,
            user=request.user
        )

        return render(
            request,
            'clientes/perfil.html',
            {
                'cliente': cliente
            }
        )

    except Exception as e:

        messages.error(
            request,
            f"Error: {str(e)}"
        )

        return redirect('tienda')


# =========================
# EDITAR PERFIL CLIENTE
# =========================
@login_required
def editar_perfil_cliente(request):

    if request.method != 'POST':
        return redirect('perfil_cliente')

    try:

        with transaction.atomic():

            cliente = get_object_or_404(
                Cliente,
                user=request.user
            )

            nombre = request.POST.get(
                'nombre',
                ''
            ).strip()

            apellido = request.POST.get(
                'apellido',
                ''
            ).strip()

            username = request.POST.get(
                'username',
                ''
            ).strip()

            # validar username
            existe = User.objects.exclude(
                id=request.user.id
            ).filter(
                username=username
            ).exists()

            if existe:

                messages.error(
                    request,
                    "El username ya existe"
                )

                return redirect('perfil_cliente')

            # actualizar user
            request.user.username = username
            request.user.first_name = nombre
            request.user.last_name = apellido

            request.user.save()

            # actualizar cliente
            cliente.nombre = nombre
            cliente.apellido = apellido

            cliente.save()

            messages.success(
                request,
                "Perfil actualizado"
            )

            return redirect('perfil_cliente')

    except Exception as e:

        messages.error(
            request,
            f"Error: {str(e)}"
        )

        return redirect('perfil_cliente')


# =========================
# AUTOCOMPLETE CLIENTES
# =========================
def buscar_clientes(request):

    term = request.GET.get('term', '')

    clientes = Cliente.objects.filter(
        nombre__icontains=term
    )[:10]

    data = [
        {
            'id': c.cliente_id,
            'text': str(c)
        }
        for c in clientes
    ]

    return JsonResponse(data, safe=False)